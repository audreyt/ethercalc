"""File processing utilities for various spreadsheet formats."""

import csv
import io
import mimetypes
from typing import Any, Dict, List, Optional, Tuple, Union
import zipfile
from pathlib import Path

import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from pydantic import BaseModel

from .exceptions import FileProcessingError


class FileTypeDetector:
    """Detects file types based on content and extension."""
    
    SUPPORTED_TYPES = {
        'csv': ['text/csv', 'application/csv'],
        'xlsx': ['application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'],
        'xls': ['application/vnd.ms-excel'],
        'ods': ['application/vnd.oasis.opendocument.spreadsheet'],
        'tsv': ['text/tab-separated-values'],
        'txt': ['text/plain'],
        'json': ['application/json']
    }
    
    @classmethod
    def detect_file_type(cls, filename: str, content: bytes) -> str:
        """Detect file type from filename and content."""
        # First try by extension
        file_path = Path(filename)
        extension = file_path.suffix.lower().lstrip('.')
        
        if extension in cls.SUPPORTED_TYPES:
            # Verify with content if possible
            if cls._verify_content_type(extension, content):
                return extension
        
        # Try by MIME type
        mime_type, _ = mimetypes.guess_type(filename)
        if mime_type:
            for file_type, mime_types in cls.SUPPORTED_TYPES.items():
                if mime_type in mime_types:
                    return file_type
        
        # Try by content magic bytes
        return cls._detect_by_content(content)
    
    @classmethod
    def _verify_content_type(cls, file_type: str, content: bytes) -> bool:
        """Verify file type by examining content."""
        if not content:
            return False
        
        # Check magic bytes
        if file_type == 'xlsx':
            # ZIP file signature (XLSX is a ZIP archive)
            return content.startswith(b'PK\x03\x04')
        elif file_type == 'xls':
            # OLE2 signature
            return content.startswith(b'\xd0\xcf\x11\xe0')
        elif file_type == 'ods':
            # ODS is also a ZIP archive
            return content.startswith(b'PK\x03\x04')
        elif file_type in ['csv', 'tsv', 'txt']:
            # Try to decode as text
            try:
                content.decode('utf-8')
                return True
            except UnicodeDecodeError:
                try:
                    content.decode('latin-1')
                    return True
                except UnicodeDecodeError:
                    return False
        
        return True  # Default to true for other types
    
    @classmethod
    def _detect_by_content(cls, content: bytes) -> str:
        """Detect file type by content analysis."""
        if not content:
            return 'txt'
        
        # Check magic bytes
        if content.startswith(b'PK\x03\x04'):
            # Could be XLSX or ODS, need to check further
            try:
                with zipfile.ZipFile(io.BytesIO(content)) as zf:
                    if 'xl/workbook.xml' in zf.namelist():
                        return 'xlsx'
                    elif 'content.xml' in zf.namelist():
                        return 'ods'
            except zipfile.BadZipFile:
                pass
            return 'xlsx'  # Default to xlsx for ZIP files
        
        elif content.startswith(b'\xd0\xcf\x11\xe0'):
            return 'xls'
        
        # Try to detect CSV by analyzing text structure
        try:
            text_content = content.decode('utf-8')
            lines = text_content.strip().split('\n')[:10]  # Check first 10 lines
            
            # Check for consistent comma-separated structure
            comma_counts = [line.count(',') for line in lines if line.strip()]
            if comma_counts and len(set(comma_counts)) <= 2:  # Allow some variation
                return 'csv'
            
            # Check for tab-separated structure
            tab_counts = [line.count('\t') for line in lines if line.strip()]
            if tab_counts and len(set(tab_counts)) <= 2:
                return 'tsv'
            
            return 'txt'
            
        except UnicodeDecodeError:
            return 'txt'  # Binary file, treat as text


class BaseProcessor:
    """Base class for file processors."""
    
    def __init__(self):
        """Initialize processor."""
        self.encoding_options = ['utf-8', 'utf-8-sig', 'latin-1', 'cp1252']
    
    def _detect_encoding(self, content: bytes) -> str:
        """Detect text encoding of file content."""
        for encoding in self.encoding_options:
            try:
                content.decode(encoding)
                return encoding
            except UnicodeDecodeError:
                continue
        return 'utf-8'  # Default fallback
    
    def _safe_decode(self, content: bytes, encoding: str = None) -> str:
        """Safely decode bytes to string."""
        if encoding is None:
            encoding = self._detect_encoding(content)
        
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            return content.decode('utf-8', errors='replace')


class CSVProcessor(BaseProcessor):
    """Processes CSV files."""
    
    def read_file(self, content: bytes, filename: str = "") -> Dict[str, List[List[str]]]:
        """
        Read CSV file and return data structure.
        
        Returns:
            Dictionary with single sheet containing rows of data
        """
        try:
            # Detect encoding and decode
            text_content = self._safe_decode(content)
            
            # Detect delimiter
            delimiter = self._detect_delimiter(text_content)
            
            # Parse CSV
            reader = csv.reader(io.StringIO(text_content), delimiter=delimiter)
            rows = list(reader)
            
            # Filter empty rows
            rows = [row for row in rows if any(cell.strip() for cell in row)]
            
            return {"Sheet1": rows}
            
        except Exception as e:
            raise FileProcessingError(f"Failed to process CSV file: {str(e)}", filename, "csv")
    
    def write_file(self, data: Dict[str, List[List[str]]], filename: str = "") -> bytes:
        """Write data to CSV format."""
        try:
            # Use the first sheet
            sheet_name = list(data.keys())[0] if data else "Sheet1"
            rows = data.get(sheet_name, [])
            
            output = io.StringIO()
            writer = csv.writer(output)
            
            for row in rows:
                # Ensure all cells are strings
                string_row = [str(cell) if cell is not None else "" for cell in row]
                writer.writerow(string_row)
            
            return output.getvalue().encode('utf-8')
            
        except Exception as e:
            raise FileProcessingError(f"Failed to write CSV file: {str(e)}", filename, "csv")
    
    def _detect_delimiter(self, content: str) -> str:
        """Detect CSV delimiter."""
        sample_lines = content[:2048]  # Check first 2KB
        
        # Count occurrences of common delimiters
        delimiters = [',', '\t', ';', '|']
        delimiter_counts = {}
        
        for delimiter in delimiters:
            # Count consistent usage across lines
            lines = sample_lines.split('\n')[:5]  # Check first 5 lines
            counts = [line.count(delimiter) for line in lines if line.strip()]
            
            if counts and max(counts) > 0:
                # Check consistency (all counts should be similar)
                avg_count = sum(counts) / len(counts)
                consistency = sum(1 for count in counts if abs(count - avg_count) <= 1)
                delimiter_counts[delimiter] = (max(counts), consistency / len(counts))
        
        if not delimiter_counts:
            return ','  # Default
        
        # Choose delimiter with highest count and consistency
        best_delimiter = max(delimiter_counts.keys(), 
                           key=lambda d: delimiter_counts[d][0] * delimiter_counts[d][1])
        
        return best_delimiter


class ExcelProcessor(BaseProcessor):
    """Processes Excel files (XLSX/XLS)."""
    
    def read_file(self, content: bytes, filename: str = "") -> Dict[str, List[List[str]]]:
        """Read Excel file and return data structure."""
        try:
            # Use openpyxl for XLSX files
            if filename.lower().endswith('.xlsx') or self._is_xlsx_content(content):
                return self._read_xlsx(content, filename)
            else:
                # Fallback to pandas for XLS files
                return self._read_xls_pandas(content, filename)
                
        except Exception as e:
            raise FileProcessingError(f"Failed to process Excel file: {str(e)}", filename, "excel")
    
    def write_file(self, data: Dict[str, List[List[str]]], filename: str = "") -> bytes:
        """Write data to Excel format."""
        try:
            workbook = openpyxl.Workbook()
            
            # Remove default sheet
            workbook.remove(workbook.active)
            
            for sheet_name, rows in data.items():
                worksheet = workbook.create_sheet(title=sheet_name[:31])  # Excel sheet name limit
                
                for row_idx, row in enumerate(rows, 1):
                    for col_idx, cell_value in enumerate(row, 1):
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        cell.value = cell_value
            
            # Save to bytes
            output = io.BytesIO()
            workbook.save(output)
            output.seek(0)
            
            return output.read()
            
        except Exception as e:
            raise FileProcessingError(f"Failed to write Excel file: {str(e)}", filename, "excel")
    
    def _is_xlsx_content(self, content: bytes) -> bool:
        """Check if content is XLSX format."""
        return content.startswith(b'PK\x03\x04')
    
    def _read_xlsx(self, content: bytes, filename: str) -> Dict[str, List[List[str]]]:
        """Read XLSX file using openpyxl."""
        workbook = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        result = {}
        
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            rows = []
            
            for row in worksheet.iter_rows(values_only=True):
                # Convert None values to empty strings and ensure all values are strings
                string_row = [str(cell) if cell is not None else "" for cell in row]
                # Remove trailing empty cells
                while string_row and not string_row[-1]:
                    string_row.pop()
                
                if string_row:  # Only add non-empty rows
                    rows.append(string_row)
            
            result[sheet_name] = rows
        
        workbook.close()
        return result
    
    def _read_xls_pandas(self, content: bytes, filename: str) -> Dict[str, List[List[str]]]:
        """Read XLS file using pandas."""
        try:
            # Read all sheets
            excel_data = pd.read_excel(io.BytesIO(content), sheet_name=None, header=None)
            result = {}
            
            for sheet_name, df in excel_data.items():
                # Convert DataFrame to list of lists
                rows = []
                for _, row in df.iterrows():
                    string_row = [str(cell) if pd.notna(cell) else "" for cell in row.values]
                    # Remove trailing empty cells
                    while string_row and not string_row[-1]:
                        string_row.pop()
                    
                    if string_row:  # Only add non-empty rows
                        rows.append(string_row)
                
                result[sheet_name] = rows
            
            return result
            
        except Exception as e:
            raise FileProcessingError(f"Failed to read XLS file with pandas: {str(e)}", filename, "xls")


class ODSProcessor(BaseProcessor):
    """Processes ODS (OpenDocument Spreadsheet) files."""
    
    def read_file(self, content: bytes, filename: str = "") -> Dict[str, List[List[str]]]:
        """Read ODS file and return data structure."""
        try:
            # Use pandas for ODS files
            excel_data = pd.read_excel(io.BytesIO(content), sheet_name=None, header=None, engine='odf')
            result = {}
            
            for sheet_name, df in excel_data.items():
                rows = []
                for _, row in df.iterrows():
                    string_row = [str(cell) if pd.notna(cell) else "" for cell in row.values]
                    # Remove trailing empty cells
                    while string_row and not string_row[-1]:
                        string_row.pop()
                    
                    if string_row:  # Only add non-empty rows
                        rows.append(string_row)
                
                result[sheet_name] = rows
            
            return result
            
        except Exception as e:
            raise FileProcessingError(f"Failed to process ODS file: {str(e)}", filename, "ods")
    
    def write_file(self, data: Dict[str, List[List[str]]], filename: str = "") -> bytes:
        """Write data to ODS format."""
        try:
            # Create a temporary file approach since direct ODS writing is complex
            import tempfile
            import os
            
            with tempfile.NamedTemporaryFile(suffix='.ods', delete=False) as tmp_file:
                temp_filename = tmp_file.name
            
            try:
                # Convert data to DataFrame and write
                with pd.ExcelWriter(temp_filename, engine='odf') as writer:
                    for sheet_name, rows in data.items():
                        if rows:
                            df = pd.DataFrame(rows)
                            df.to_excel(writer, sheet_name=sheet_name[:31], index=False, header=False)
                
                # Read back the file
                with open(temp_filename, 'rb') as f:
                    result = f.read()
                
                return result
                
            finally:
                # Clean up temp file
                if os.path.exists(temp_filename):
                    os.unlink(temp_filename)
            
        except Exception as e:
            raise FileProcessingError(f"Failed to write ODS file: {str(e)}", filename, "ods")


class FileHandler:
    """Main file handler that delegates to specific processors."""
    
    def __init__(self):
        """Initialize file handler with processors."""
        self.processors = {
            'csv': CSVProcessor(),
            'tsv': CSVProcessor(),  # TSV uses same processor as CSV
            'xlsx': ExcelProcessor(),
            'xls': ExcelProcessor(),
            'ods': ODSProcessor(),
        }
        self.detector = FileTypeDetector()
    
    def read_file(self, content: bytes, filename: str) -> Dict[str, List[List[str]]]:
        """
        Read file content and return standardized data structure.
        
        Args:
            content: File content as bytes
            filename: Original filename
            
        Returns:
            Dictionary mapping sheet names to lists of rows (list of cells)
        """
        try:
            # Detect file type
            file_type = self.detector.detect_file_type(filename, content)
            
            if file_type not in self.processors:
                raise FileProcessingError(f"Unsupported file type: {file_type}", filename, file_type)
            
            # Process file
            processor = self.processors[file_type]
            return processor.read_file(content, filename)
            
        except FileProcessingError:
            raise
        except Exception as e:
            raise FileProcessingError(f"Unexpected error processing file: {str(e)}", filename)
    
    def write_file(self, data: Dict[str, List[List[str]]], filename: str, 
                  file_type: str = None) -> bytes:
        """
        Write data to specified file format.
        
        Args:
            data: Data structure (sheet_name -> rows -> cells)
            filename: Target filename
            file_type: Target file type (auto-detected from filename if not provided)
            
        Returns:
            File content as bytes
        """
        try:
            if file_type is None:
                # Detect from filename extension
                file_path = Path(filename)
                file_type = file_path.suffix.lower().lstrip('.')
            
            if file_type not in self.processors:
                raise FileProcessingError(f"Unsupported output file type: {file_type}", filename, file_type)
            
            processor = self.processors[file_type]
            return processor.write_file(data, filename)
            
        except FileProcessingError:
            raise
        except Exception as e:
            raise FileProcessingError(f"Unexpected error writing file: {str(e)}", filename, file_type)
    
    def get_supported_types(self) -> List[str]:
        """Get list of supported file types."""
        return list(self.processors.keys())
    
    def validate_file(self, content: bytes, filename: str) -> Dict[str, Any]:
        """
        Validate file and return information about it.
        
        Returns:
            Dictionary with validation results and file info
        """
        try:
            file_type = self.detector.detect_file_type(filename, content)
            
            # Basic validation
            if not content:
                return {
                    'valid': False,
                    'error': 'File is empty',
                    'file_type': file_type,
                    'size': 0
                }
            
            if file_type not in self.processors:
                return {
                    'valid': False,
                    'error': f'Unsupported file type: {file_type}',
                    'file_type': file_type,
                    'size': len(content)
                }
            
            # Try to read the file to validate structure
            try:
                data = self.read_file(content, filename)
                
                # Count sheets and rows
                sheet_count = len(data)
                total_rows = sum(len(rows) for rows in data.values())
                max_cols = max((len(row) for rows in data.values() for row in rows), default=0)
                
                return {
                    'valid': True,
                    'file_type': file_type,
                    'size': len(content),
                    'sheets': sheet_count,
                    'total_rows': total_rows,
                    'max_columns': max_cols,
                    'sheet_names': list(data.keys())
                }
                
            except Exception as e:
                return {
                    'valid': False,
                    'error': f'File structure validation failed: {str(e)}',
                    'file_type': file_type,
                    'size': len(content)
                }
                
        except Exception as e:
            return {
                'valid': False,
                'error': f'File validation failed: {str(e)}',
                'file_type': 'unknown',
                'size': len(content) if content else 0
            }


class ImportOptions(BaseModel):
    """Options for importing data into spreadsheet."""
    delimiter: str = ","
    encoding: str = "utf-8"
    has_header: bool = True
    start_cell: str = "A1"
    sheet_name: Optional[str] = None
    max_rows: Optional[int] = None
    max_cols: Optional[int] = None


class ExportOptions(BaseModel):
    """Options for exporting data from spreadsheet."""
    range_ref: Optional[str] = None
    delimiter: str = ","
    include_header: bool = True
    include_formatting: bool = False
    sheet_name: Optional[str] = None


async def import_csv(csv_content: str, spreadsheet, options: ImportOptions, user_id: Optional[str] = None) -> Dict[str, Any]:
    """Import CSV content into spreadsheet."""
    try:
        # Parse CSV
        reader = csv.reader(io.StringIO(csv_content), delimiter=options.delimiter)
        rows = list(reader)
        
        if not rows:
            return {}
        
        # Skip header if specified
        start_row_idx = 1 if options.has_header else 0
        data_rows = rows[start_row_idx:]
        
        # Apply limits
        if options.max_rows:
            data_rows = data_rows[:options.max_rows]
        
        if options.max_cols:
            data_rows = [row[:options.max_cols] for row in data_rows]
        
        # Parse start cell
        start_col, start_row = spreadsheet._parse_cell_ref(options.start_cell)
        
        # Import data
        imported_cells = {}
        for row_idx, row_data in enumerate(data_rows):
            for col_idx, cell_value in enumerate(row_data):
                if cell_value.strip():  # Only import non-empty cells
                    cell_ref = spreadsheet._create_cell_ref(start_col + col_idx, start_row + row_idx)
                    imported_cells[cell_ref] = spreadsheet.set_cell(cell_ref, cell_value, user_id=user_id)
        
        return imported_cells
    except Exception as e:
        raise FileProcessingError(f"Failed to import CSV: {str(e)}")


async def import_xlsx(xlsx_content: io.BytesIO, spreadsheet, options: ImportOptions, user_id: Optional[str] = None) -> Dict[str, Any]:
    """Import XLSX content into spreadsheet."""
    try:
        # Load workbook
        workbook = openpyxl.load_workbook(xlsx_content, data_only=True)
        
        # Select sheet
        if options.sheet_name and options.sheet_name in workbook.sheetnames:
            worksheet = workbook[options.sheet_name]
        else:
            worksheet = workbook.active
        
        # Parse start cell
        start_col, start_row = spreadsheet._parse_cell_ref(options.start_cell)
        
        # Get data
        data_rows = []
        for row in worksheet.iter_rows(values_only=True):
            row_data = [str(cell) if cell is not None else "" for cell in row]
            # Remove trailing empty cells
            while row_data and not row_data[-1]:
                row_data.pop()
            if row_data:  # Only keep non-empty rows
                data_rows.append(row_data)
        
        # Skip header if specified
        if options.has_header and data_rows:
            data_rows = data_rows[1:]
        
        # Apply limits
        if options.max_rows:
            data_rows = data_rows[:options.max_rows]
        
        if options.max_cols:
            data_rows = [row[:options.max_cols] for row in data_rows]
        
        # Import data
        imported_cells = {}
        for row_idx, row_data in enumerate(data_rows):
            for col_idx, cell_value in enumerate(row_data):
                if cell_value.strip():  # Only import non-empty cells
                    cell_ref = spreadsheet._create_cell_ref(start_col + col_idx, start_row + row_idx)
                    imported_cells[cell_ref] = spreadsheet.set_cell(cell_ref, cell_value, user_id=user_id)
        
        workbook.close()
        return imported_cells
    except Exception as e:
        raise FileProcessingError(f"Failed to import XLSX: {str(e)}")


async def export_csv(spreadsheet, options: ExportOptions) -> str:
    """Export spreadsheet to CSV format."""
    try:
        # Get cells to export
        if options.range_ref:
            cells = spreadsheet.get_range(options.range_ref)
        else:
            cells = spreadsheet.cells
        
        if not cells:
            return ""
        
        # Determine bounds
        cell_refs = list(cells.keys())
        cols = []
        rows = []
        
        for ref in cell_refs:
            col, row = spreadsheet._parse_cell_ref(ref)
            cols.append(col)
            rows.append(row)
        
        min_col, max_col = min(cols), max(cols)
        min_row, max_row = min(rows), max(rows)
        
        # Build CSV content
        output = io.StringIO()
        writer = csv.writer(output, delimiter=options.delimiter)
        
        for row in range(min_row, max_row + 1):
            csv_row = []
            for col in range(min_col, max_col + 1):
                cell_ref = spreadsheet._create_cell_ref(col, row)
                cell = cells.get(cell_ref)
                value = ""
                if cell:
                    if cell.cell_type.value == "formula" and cell.value is not None:
                        value = str(cell.value)  # Use calculated value
                    elif cell.value is not None:
                        value = str(cell.value)
                csv_row.append(value)
            
            # Remove trailing empty cells
            while csv_row and not csv_row[-1]:
                csv_row.pop()
            
            if csv_row:  # Only write non-empty rows
                writer.writerow(csv_row)
        
        return output.getvalue()
    except Exception as e:
        raise FileProcessingError(f"Failed to export CSV: {str(e)}")


async def export_xlsx(spreadsheet, options: ExportOptions) -> io.BytesIO:
    """Export spreadsheet to XLSX format."""
    try:
        # Create workbook
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = spreadsheet.metadata.title[:31]  # Excel title limit
        
        # Get cells to export
        if options.range_ref:
            cells = spreadsheet.get_range(options.range_ref)
        else:
            cells = spreadsheet.cells
        
        if cells:
            # Determine bounds
            cell_refs = list(cells.keys())
            cols = []
            rows = []
            
            for ref in cell_refs:
                col, row = spreadsheet._parse_cell_ref(ref)
                cols.append(col)
                rows.append(row)
            
            min_col, max_col = min(cols), max(cols)
            min_row, max_row = min(rows), max(rows)
            
            # Write data to worksheet
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    cell_ref = spreadsheet._create_cell_ref(col, row)
                    cell_data = cells.get(cell_ref)
                    
                    if cell_data and cell_data.value is not None:
                        excel_cell = worksheet.cell(row=row + 1, column=col + 1)
                        
                        if cell_data.cell_type.value == "formula" and cell_data.value is not None:
                            excel_cell.value = cell_data.value  # Use calculated value
                        else:
                            excel_cell.value = cell_data.value
                        
                        # Apply formatting if requested
                        if options.include_formatting and cell_data.format:
                            format_data = cell_data.format
                            if format_data.font_bold:
                                excel_cell.font = openpyxl.styles.Font(bold=True)
                            if format_data.background_color:
                                fill = openpyxl.styles.PatternFill(
                                    start_color=format_data.background_color.lstrip('#'),
                                    end_color=format_data.background_color.lstrip('#'),
                                    fill_type="solid"
                                )
                                excel_cell.fill = fill
        
        # Save to bytes
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        
        workbook.close()
        return output
    except Exception as e:
        raise FileProcessingError(f"Failed to export XLSX: {str(e)}")